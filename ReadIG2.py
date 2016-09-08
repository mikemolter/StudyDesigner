import openpyxl
from py2neo import Graph

wb = openpyxl.load_workbook('Initial SDTM 1.4 IG 3.2.xlsx')
IG_datasets = wb.get_sheet_by_name('SDTMIG 3.2 Domains')
IG_Methods = wb.get_sheet_by_name('SDTM 1.4 Mapping Specs')

graph = Graph()
tx = graph.cypher.begin()

tx.append("CREATE (ig:IG {model:'SDTM', version:'3.2'}) WITH ig MATCH (m:Model {name:'SDTM',version:'1.4'}) CREATE (ig)-[r:InstantiateModel]->(m)")

# Import domain level metadata
propertynames = ['Name', 'Label', 'Class', 'Structure', 'Purpose', 'Reference', 'Repeating', 'URL']
for row in range(2,IG_datasets.max_row+1):
    statement = "MERGE (IGD:IGDomains {"
    firstproperty = True
    for col in range(len(propertynames)):
        if IG_datasets.cell(row=row,column=col+1).value:
            if not firstproperty:
                statement=statement+', '
            statement=statement+propertynames[col]+': "'+str(IG_datasets.cell(row=row,column=col+1).value)+'" '
            firstproperty = False
    statement=statement+'}) WITH IGD MATCH (ig:IG {model:"SDTM"}) MERGE (ig)-[r:ContainsDomain]->(IGD)'
    tx.append(statement)

    domain = IG_datasets.cell(row=row,column=1).value
    IG_vars = wb.get_sheet_by_name('SDTMIG 3.2 Items '+domain+' V2')
    defpropsIG = ['Name','ModelOID','Label','Role','Core']
    defpropsML=['SASType','Length','DataType','CodeList','Dictionary','Origin','VLMFlag']
    refpropsIG = ['Order','MapSpecOID','Key','Mandatory']

    # Create domain-specific methods
    for rowMT in range(2,IG_Methods.max_row+1):
        MethodOID = str(IG_Methods.cell(row=rowMT,column=1).value) %{"domain":domain}
        Method    = str(IG_Methods.cell(row=rowMT,column=2).value) %{"domain":domain}
        statement = 'merge (a:Method {SpecOID:"'+MethodOID+'",MapSpec:"'+Method+'"})'
        tx.append(statement)

    # Now create Item nodes
    for rowIG in range(2,IG_vars.max_row+1):
    #for rowIG in [2,5]:
        ModelOID=str(IG_vars.cell(row=rowIG,column=2).value)
        IGOID=ModelOID %{"domain":domain,"studyname":"%(studyname)s"}
        statement='match (a:ModelItem {OID:"'+ModelOID+'"}) with a merge (b:IGItem {OID:"'+IGOID+'"}) on create set '
        for colM in range(len(defpropsML)):
            statement=statement+'b.'+defpropsML[colM]+'= a.'+defpropsML[colM]+', '
        for colI in range(len(defpropsIG)):
            if colI != 1:
                statement=statement+'b.'+defpropsIG[colI]+'="'+str(IG_vars.cell(row=rowIG,column=colI+1).value)+'"'
                if colI != len(defpropsIG)-1:
                    statement=statement+', '
        #print statement
        tx.append(statement)

        # Create relationships
        firstproperty=True
        statement='match (a:IGDomains {Name:"'+domain+'"}),(b:IGItem {OID:"'+IGOID+'"}) create (a)-[:ContainsIGItem {'
        for colIR in range(len(refpropsIG)):
            if IG_vars.cell(row=rowIG,column=colIR+6).value:
                if not firstproperty:
                    statement=statement+', '
                statement=statement+refpropsIG[colIR]+':"'+str(IG_vars.cell(row=rowIG,column=colIR+6).value)+'"'
                firstproperty=False
        statement=statement+'}]->(b)'
        #print statement
        tx.append(statement)
        
        # Now create relationships to codelists and Methods
        if str(IG_vars.cell(row=rowIG,column=7).value):
            statement='match (b:Method {SpecOID:"'+str(IG_vars.cell(row=rowIG,column=7).value)+'"}),(c:IGItem {OID:"'+IGOID+'"}) create (c)-[:UsesSpec]->(b)'
            tx.append(statement)
            #print statement
        statement='match (a:ModelItem {OID:"'+ModelOID+'"})-[:UsesCodeList]->(b:CodeList),(c:IGItem {OID:"'+IGOID+'"}) with b,c create (c)-[:UsesCodeList]->(b)'
        tx.append(statement)

tx.commit()
