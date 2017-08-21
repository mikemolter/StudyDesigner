import openpyxl
from py2neo import Graph
from py2neo.packages.httpstream import http
http.socket_timeout = 9999
CTVersion='2017Q2'
wbsdtm='SDTM Terminology Q2.xlsx'
wbadam='ADaM Terminology.xlsx'
sheetsdtm='SDTM Terminology 2017-06-30'
sheetadam='ADaM Terminology 2017-03-31'

wbs = openpyxl.load_workbook(wbsdtm)
cts = wbs.get_sheet_by_name(sheetsdtm)
wba = openpyxl.load_workbook(wbadam)
cta = wba.get_sheet_by_name(sheetadam)
ct  = [cts,cta]

graph = Graph('http://neo4j:letsgowings@10.0.0.10:7474/db/data/')
tx = graph.cypher.begin()

tx.append('CREATE (ct:CT {version: "'+CTVersion+'"}) with ct ')

clipropnames=['AliasName','CodedValue','Synonym','Definition','Decode']
clicolumns=[1,5,6,7,8]
clpropnames=['AliasName','Extensible','Name','OID','Synonym','Definition','PreferredTerm']
clcolumns=[1,3,4,5,6,7,8]

for c in ct:
    for row in range(2,ct.max_row+1):
        firstproperty=True
        if c.cell(row=row,column=2).value:
            statement='MERGE (cli:CodeListItem {'
            for col in range(len(clipropnames)):
                if not firstproperty:
                    statement=statement+', '
                statement=statement+clipropnames[col]+': "'+str(c.cell(row=row,column=clicolumns[col]).value)+'" '
                firstproperty=False
            statement=statement+'}) WITH cli MATCH (cl:CodeList {Aliasname: "'+str(c.cell(row=row,column=2).value)+'"}) CREATE (cl)-[r:ContainsCodeListItem]->(cli)'
        else:
            statement='CREATE (ct)-[:ContainsCodeList {Child:0}]->(cl:CodeList {'
            for col in range(len(clpropnames)):
                if not firstproperty:
                    statement=statement+', '
                statement=statement+clpropnames[col]+': "'+str(c.cell(row=row,column=clcolumns[col]).value)+'"'
                firstproperty=False
            statement=statement+'})'
        tx.append(statement)
tx.commit()

# Changes to make
# 4) Add YESONLY BasedOn relationship with child=1
# 5) Use Decode flag to differentiate which codelists use enumerateditems and which don't
