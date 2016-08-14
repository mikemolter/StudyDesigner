import openpyxl
from py2neo import Graph

wb = openpyxl.load_workbook('Initial SDTM 1.4 IG 3.2.xlsx')
IG_TEST = wb.get_sheet_by_name('Test')

graph = Graph()
tx = graph.cypher.begin()

propertynames = ['ID','PROP1','PROP2']
for row in range(2,IG_TEST.max_row+1):
    statement='CREATE (:TestServer {'
    firstproperty=True
    for col in range(3):
        if not firstproperty:
            statement=statement+', '
        firstproperty=False
        statement=statement+propertynames[col]+':"'+str(IG_TEST.cell(row=row,column=col+1).value)+'" '
    statement=statement+'}) '
    tx.append(statement)
tx.commit()
