from django.shortcuts import render
from django.http import HttpResponse
from py2neo import Graph
import openpyxl
from datetime import datetime
from lxml import etree


#graph = Graph('http://neo4j:letsgowings@10.0.0.10:7474/db/data/')
graph = Graph('http://neo4j:letsgowings@localhost:7474/db/data/')

# Create your views here.
def index(request):
	versions=graph.cypher.execute('match (a:CT) return a.version as version')
	currentstudies=graph.cypher.execute('match (a:Study) return a.Name as study')
	return render(request,'CTVag/index.html',{'Versions':versions, 'Studies':currentstudies})

def dbcodelists(request):
	global chosenversion
	global indexchoice
	global study
	global ScopeNode

	# Determine which choice was made
	indexchoice = request.GET['indexchoice'] 
	# If choice was to create child codelists at the global level ...
	if indexchoice == 'global':
		# Determine which NCI version
		chosenversion=request.GET['CTVersions']
		ScopeNode = 'CT {version:"'+chosenversion+'"}'
		QueryDB(ScopeNode)
		# Render a page with all global codelists of the chosen version
		return render(request,'CTVag/dbcodelists.html',{'CodeLists':DBQuery,'MyVersion':chosenversion})

	# Otherwise if choice was study-level ...
	else:
		# Create a new study
		if indexchoice == 'newstudy':
			study = request.GET['newstudyname']
			chosenversion=request.GET['CTVersions']
			# Create a study node that points back to the chosen CT version with a Scope relationship
			tx=graph.cypher.begin()
			tx.append('match (a:CT {version:"'+chosenversion+'"}) with a create (a)<-[:Scope]-(b:Study {Name:"'+study+'", Purpose: "CT Only"})')
			tx.commit()
		else:
			study = request.GET['currentstudyname']
			CVRL = graph.cypher.execute('match (a:CT)<-[:Scope]-(b:Study {Name:"'+study+'"}) return a.version')
			chosenversion = CVRL[0][0]

		ScopeNode = 'Study {Name:"'+study+'", Purpose: "CT Only"}'
		QueryDB(ScopeNode)
		# Render a page with all codelists for the chosen/created study
		return render(request,'CTVag/studycodelists.html',{'Codelists':DBQuery,'Studyname':study})

def displayNCI(request):
	QueryDB(ScopeNode)
	return render(request,'CTVag/dbcodelists.html',{'CodeLists':DBQuery,'MyVersion':chosenversion})

def displayStudy(request):
	print 'From displayStudy, SCOPENODE: ' + ScopeNode
	QueryDB(ScopeNode)
	return render(request,'CTVag/studycodelists.html',{'Codelists':DBQuery,'Studyname':study})

def FromStudy(request):
	studychoice = request.GET['studychoice']
	print 'From FromStudy, SCOPENODE: ' + ScopeNode
	
	if studychoice == "Add an NCI codelist":
		QueryDB('CT {version:"'+chosenversion+'"}')
		return render(request,'CTVag/AddNCI.html',{'CodeLists':DBQuery,'MyVersion':chosenversion})

	elif studychoice == "Define a sponsor codelist":
		return render (request,'CTVag/NewCodeList1.html')

	elif studychoice == "Home":
		versions=graph.cypher.execute('match (a:CT) return a.version as version')
		currentstudies=graph.cypher.execute('match (a:Study) return a.Name as study')
		return render(request,'CTVag/index.html',{'Versions':versions, 'Studies':currentstudies})

	else:
		StudyCodeLists=graph.cypher.execute('match (a:'+ScopeNode+')-[:ContainsCodeList]->(b:CodeList)-[:ContainsCodeListItem]->(c:CodeListItem) return \
			b.OID as OID,b.Name as Name,b.AliasName as CLAlias,b.DataType as DataType,c.CodedValue as CodedValue,c.AliasName as TermAlias,c.Decode as Decode order by b.OID,c.CodedValue')

		if studychoice == "Output Excel":
			wb = openpyxl.Workbook()
			sheet = wb.active
			sheet.title = "Codelists"
			sheet['A1'] = 'ID'
			sheet['B1'] = 'Name'
			sheet['C1'] = 'NCI Codelist Code'
			sheet['D1'] = 'Data Type'
			sheet['E1'] = 'Order'
			sheet['F1'] = 'Term'
			sheet['G1'] = 'NCI Term Code'
			sheet['H1'] = 'Decoded Value'
			row = 1
			for x1 in StudyCodeLists:
				order=1
				row = row+1
				sheet.cell(row=row,column=1).value=x1[0]
				sheet.cell(row=row,column=2).value=x1[1]
				sheet.cell(row=row,column=3).value=x1[2]
				sheet.cell(row=row,column=4).value=x1[3]
				sheet.cell(row=row,column=5).value=order
				sheet.cell(row=row,column=6).value=x1[4]
				sheet.cell(row=row,column=7).value=x1[5]
				sheet.cell(row=row,column=8).value=x1[6]
				order = order + 1
			filename = 'ct-'+datetime.now().isoformat()+'.xlsx'
			wb.save(filename)

		elif studychoice == 'Output XML':
			root=etree.Element('ODM')
			lastOID=''
			xmlns='http://www.w3.org/XML/1998/namespace'
			ttatts={etree.QName(xmlns,'lang'):'en'}
			for x in StudyCodeLists:
				if x['OID'] != lastOID:
					CL=etree.SubElement(root,'CodeList',OID=x['OID'],Name=x['Name'],DataType=x['DataType'])
					if x['CLAlias']:
						aliasCL=etree.SubElement(CL,'Alias',Name=x['CLAlias'],Context='nci:ExtCodeID')
				lastOID=x['OID']

				if x['Decode']:
					cli=etree.SubElement(CL,'CodeListItem',CodedValue=x['CodedValue'])
					dec=etree.SubElement(cli,'Decode')
					tt=etree.SubElement(dec,'TranslatedText',ttatts)
					tt.text=x['Decode']
					if x['TermAlias']:
						aliasterm=etree.SubElement(cli,'Alias',Name=x['TermAlias'],Context='nci:ExtCodeID')

				else:
					enum=etree.SubElement(CL,'EnumeratedItem',CodedValue=x['CodedValue'])
					if x['TermAlias']:
						aliasterm=etree.SubElement(enum,'Alias',Name=x['TermAlias'],Context='nci:ExtCodeID')

			filename = 'ct-'+datetime.now().isoformat()+'.xml'
			doc=etree.ElementTree(root)
			doc.write(filename,pretty_print=True)

		return render(request,'CTVag/summaryoutput.html',{'filename':filename})

def QueryDB(node):
	global DBQuery
	# DBQuery=graph.cypher.execute('match (a:'+ScopeNodeQuery+')-[r1:ContainsCodeList]->(b:CodeList) \
	# 	with collect({Name:b.Name, Parent:b.OID,OID:b.OID,AliasName:b.AliasName,child:b.child,DataType:b.DataType}) as rows \
	# 	match (a:CodeList)-[r2:BasedOn]->(b:CodeList)<-[r3:ContainsCodeList]-(c:'+ScopeNodeQuery+') \
	# 	with rows+collect({Name:a.Name, Parent:b.OID,OID:a.OID,AliasName:a.AliasName,child:a.child,DataType:b.DataType}) as allrows \
	# 	unwind allrows as row with row.Name as Name,row.Parent as ParentOID,row.OID as OID,row.AliasName as AliasName,row.child as child, \
	# 	row.DataType as DataType return Name,ParentOID,OID,AliasName,child,DataType order by ParentOID,child')

	# DBQuery=graph.cypher.execute('match (a:'+node+')-[r1:ContainsCodeList]->(b:CodeList) \
	# 	with collect({Name:b.Name, Parent:b.OID, OID:b.OID, AliasName:b.AliasName, child:b.Child}) as rows \
	# 	match (a:CodeList)-[:Scope]->(:'+node+'),(a)-[:BasedOn]->(b:CodeList)<-[:ContainsCodeList]-(c:'+node+') \
	# 	with rows+collect({Name:a.Name, Parent:b.OID, OID:a.OID, AliasName:a.AliasName, child:a.Child}) as allrows \
	# 	unwind allrows as row with row.Name as Name,row.Parent as ParentOID,row.OID as OID,row.AliasName as AliasName,row.child as child \
	# 	return Name,ParentOID,OID,AliasName,child order by ParentOID,child')
	DBQuery=graph.cypher.execute('match (a:'+node+')-[r1:ContainsCodeList]->(b:CodeList) where r1.Child=0\
		with collect({Name:b.Name, Parent:b.OID, OID:b.OID, AliasName:b.AliasName, child:r1.Child}) as rows \
		match (a:'+node+')-[r1:ContainsCodeList]->(b:CodeList)-[:BasedOn]->(c:CodeList) where r1.Child>0 \
		with rows+collect({Name:b.Name, Parent:c.OID, OID:b.OID, AliasName:b.AliasName, child:r1.Child}) as allrows \
		unwind allrows as row with row.Name as Name,row.Parent as ParentOID,row.OID as OID,row.AliasName as AliasName,row.child as child \
		return Name,ParentOID,OID,AliasName,child order by ParentOID,child')

	if not DBQuery:
		DBQuery=graph.cypher.execute('match (a:'+node+')-[r1:ContainsCodeList]->(b:CodeList) return b.Name as Name,b.OID as ParentOID, \
			b.OID as OID,b.AliasName as AliasName,r1.Child as child order by b.OID,r1.Child')

def ChildrenCL(request):
	# Determine which button was clicked
	for k,v in request.POST.iteritems():
		
		# A codelist button was clicked to create a child
		if k[0:2] == 's_':

			# Determine if the chosen codelist is a parent.  
			parent=graph.cypher.execute('match (a:'+ScopeNode+')-[r:ContainsCodeList]->(b:CodeList {OID:"'+k[2:]+'"}) return r.Child')
			# If so, ...
			if parent[0][0] == 0:
				# Determine number of children
				countchildren=graph.cypher.execute('match (a:'+ScopeNode+')-[:ContainsCodeList]-> \
					(b:CodeList {OID:"'+k[2:]+'"})<-[:BasedOn]-(c:CodeList)<-[:ContainsCodeList]-(d:'+ScopeNode+') return count(c) as count')

				# Get the codelist items of the chosen codelist
				codelistitems=graph.cypher.execute('match (a:'+ScopeNode+')-[:ContainsCodeList]- \
					(b:CodeList {OID:"'+k[2:]+'"})-[:ContainsCodeListItem]->(d:CodeListItem) \
					return b.OID as parentOID,b.Extensible as parentext,b.Name as parentname,b.AliasName as parentalias,b.DataType as DataType, \
					d.CodedValue as codedvalue,d.Decode as decode')

			# Otherwise, if the chosen codelist is already a child...
			else:
				# Determine the number of children the selected codelist's parent has
				countchildren=graph.cypher.execute('match (a:'+ScopeNode+')-[:ContainsCodeList]-> \
					(b:CodeList {OID:"'+k[2:]+'"})-[:BasedOn]->(c:CodeList) with c match (c)<-[:BasedOn]-(d:CodeList)<-[:ContainsCodeList]-(e:'+ScopeNode+')  return count(d) as count')

				codelistitems=graph.cypher.execute('match (a:'+ScopeNode+')-[:ContainsCodeList]-> \
					(b:CodeList {OID:"'+k[2:]+'"})-[:BasedOn]->(c:CodeList) , (b) \
					-[:ContainsCodeListItem] -> (e:CodeListItem) return c.OID as parentOID,c.Extensible as parentext,c.Name as parentname, \
					c.AliasName as parentalias,c.DataType as DataType,e.CodedValue as codedvalue,e.Decode as decode')

			nextchild=countchildren[0][0]+1
			return render(request,'CTVag/childrencl.html',{'Items':codelistitems,'NextChild':nextchild})

		elif k[0:2] == 'r_':
			tx = graph.cypher.begin()
			statement = 'match (a:'+ScopeNode+')-[r:ContainsCodeList]->(b:CodeList {OID:"'+k[2:]+'"}) delete r'
			print 'STATEMENT: '
			print statement 
			tx.append(statement)
			tx.commit()
			QueryDB(ScopeNode)
			return render(request,'CTVag/studycodelists.html',{'Codelists':DBQuery,'Studyname':study})

def NewChild(request):
	OID=request.POST['OID']
	Name=request.POST['Name']
	datatype=request.POST['datatype']
	if 'decodeflag' in request.POST:
		decodeflag=request.POST['decodeflag']
	else:
		decodeflag=''
	return render(request,'CTVag/Newchildrencl.html',{'OID':OID,'Name':Name,'datatype':datatype,'decodeflag':decodeflag})

def DBChangeFromChild(request):
	if request.POST['submit'] == 'Save':
		nextchildstr=str(request.POST['NextChild'])

		# Create a CodeList node
		tx=graph.cypher.begin()

		newname=request.POST["ParentName"]
		if request.POST['newname']:
			newname=newname+' - '+request.POST['newname']

		# Create OID values of the form ParentOIDn_(NCI|Study) where n is the next child number
		if request.POST['ParentOID'].rfind('_') > 0:
			OID = request.POST['ParentOID'].rpartition('_')[0]+nextchildstr
			poid=request.POST['ParentOID']
			rp=poid.rpartition('_')
			print 'POID: '
			print poid
			print 'RP: '
			print rp
		else:
			OID=request.POST['ParentOID']+nextchildstr

		if indexchoice == 'global':
			OID = OID+'_NCI'
		else:
			OID = OID+'_'+study 

		# Create the new codelist, attach it to the parent with BasedOn relationship, and to the study or NCI node with a ContainsCodeList relationship that has a numeric Child property
		# statement = 'match (a:'+ScopeNode+')-[r:ContainsCodeList]->(b:CodeList {OID:"'+request.POST['ParentOID']+'"}) with a,b \
		# 	create (c:CodeList {OID:"'+OID+'", DataType:"'+request.POST['DataType']+'", Name:"'+newname+ \
		# 	'", AliasName: "'+request.POST['Alias']+'", Extensible: "'+request.POST['Extensible']+'"})- \
		# 	[r2:BasedOn]->(b),(a)-[:ContainsCodeList {Child:'+nextchildstr+'}]->(c) with c '

		statement = 'match (a:'+ScopeNode+'),(b:CodeList {OID:"'+request.POST['ParentOID']+'"}) with a,b \
			create (c:CodeList {OID:"'+OID+'", DataType:"'+request.POST['DataType']+'", Name:"'+newname+ \
			'", AliasName: "'+request.POST['Alias']+'", Extensible: "'+request.POST['Extensible']+'"})- \
			[r2:BasedOn]->(b),(a)-[:ContainsCodeList {Child:'+nextchildstr+'}]->(c) with c '

		# Create the CodeListItem nodes and relationships to the CodeList node
		firstitem=True 
		for k,v in request.POST.iteritems():
			# Relationships to already-existing codelist items
			if k[0:2] == 'c_':
				if not firstitem:
					statement=statement+' with c '
				firstitem=False
				# statement=statement+'match (:'+ScopeNode+')-[:ContainsCodeList]->(:CodeList {OID:"'+request.POST['ParentOID']+'"}) \
				# 	-[:ContainsCodeListItem]->(d:CodeListItem {CodedValue:"'+k[2:]+'"}) with c,d create (c)-[:ContainsCodeListItem]->(d) '
				statement=statement+'match (:CodeList {OID:"'+request.POST['ParentOID']+'"}) \
					-[:ContainsCodeListItem]->(d:CodeListItem {CodedValue:"'+k[2:]+'"}) with c,d create (c)-[:ContainsCodeListItem]->(d) '
			# Creation of and Relationships to extension items
			if k[0:3] == 'ec_':
				extval=k[3:]
				if not firstitem:
					statement=statement+' with c '
				firstitem=False
				statement=statement+'create (c)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+extval+'", '
				if request.POST['e_'+extval]:
					statement=statement+'Decode:"'+request.POST['e_'+extval]+'", '
				statement=statement+'extendedValue:"Yes"}) '

		tx.append(statement)
		tx.commit()

		return render(request,'CTVag/summarya.html',{'NewName':newname,'OID':OID,'IndexChoice':indexchoice})

	else:
		if indexchoice == 'global':
			return render(request,'CTVag/dbcodelists.html',{'CodeLists':DBQuery,'MyVersion':chosenversion})
		else:
			return render(request,'CTVag/studycodelists.html',{'Codelists':DBQuery,'Studyname':study})

def DBChangeFromNew(request):
	print 'FROM DBCHANGEFROMNEW'
	if request.POST['submit'] == 'Save':
		statement = 'match (a:'+ScopeNode+') with a create (a)-[:ContainsCodeList {Child:0}]->(b:CodeList {OID:"'+request.POST['OID']+'", \
			DataType:"'+request.POST['datatype']+'", Name:"'+request.POST["Name"]+'"}) with b '

		firstitem=True
		for k,v in request.POST.iteritems():
			if k[0:2] == 'e_':
				if not firstitem:
					statement=statement+' with b '
				firstitem=False
				statement=statement+'create (b)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+k[2:]+'"'
				if v:
					statement=statement+', Decode:"'+v+'"'
				statement=statement+'})'

		tx=graph.cypher.begin()
		tx.append(statement)
		tx.commit()

		return render(request,'CTVag/summarya.html',{'OID':request.POST['OID'],'NewName':request.POST['Name'],'nextchildstr':'','IndexChoice':indexchoice})

	else:
		return render(request,'CTVag/studycodelists.html',{'Codelists':DBQuery,'Studyname':study})


def DBChangeFromAddNCI(request):
	if request.POST['submit'] == 'Submit':
		print "SUBMIT: " + request.POST['submit']
		tx=graph.cypher.begin()
		for k,v in request.POST.iteritems():
			print "KV: "+k+" "+v
			if k[0:2] == 'a_':
				# Connect the study to the codelist
				statement = 'match (a:'+ScopeNode+'), (b:CT {version:"'+chosenversion+'"})-[:ContainsCodeList]->(c:CodeList {OID:"'+k[2:]+'"}) with a,c create (a)-[:ContainsCodeList {Child:0}]->(c)'
				tx.append(statement)
				print "STATEMENT: "+statement
		
		tx.commit()

	# Whether submitted or canceled, go back to the list of study codelists
	StudyQuery = graph.cypher.execute('match (a:'+ScopeNode+')-[:ContainsCodeList]->(b:CodeList) return b.OID as OID,b.Name as Name')
	return render(request,'CTVag/studycodelists.html',{'Codelists':StudyQuery,'Studyname':study})



