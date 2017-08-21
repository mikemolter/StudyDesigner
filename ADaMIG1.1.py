import openpyxl
from py2neo import Graph

modelname='ADaM'
modelversion='2.1'
stdname='ADaMIG1.1'

wb = openpyxl.load_workbook('adam-2-1-draft.xlsx')
IG_datasets = wb.get_sheet_by_name('ADaMIG 1.1 Dataset')
IG_vars = wb.get_sheet_by_name('ADaMIG 1.1 Variable')

#graph = Graph("http://neo4j:letsgowings@10.0.0.10:7474/db/data/")
graph = Graph('http://neo4j:letsgowings@localhost:7474/db/data/')

tx = graph.cypher.begin()

# Create the standard node and attach to the model
tx.append("match (m:Model {name:'ADaM', version:'2.1'}) create (s:Standard {name:'"+stdname+"'})-[:BasedOn]->(m)")

# Import dataset level metadata, create the itemgroupdef nodes, and attach to the standard
propertynames=['Name','Label','Class','Structure','Purpose','Reference','Repeating']

for row in range(2,IG_datasets.max_row+1):
	statement = "match (s:Standard {name:'ADaMIG', version:'1.1'}),(c:ItemGroupDef {Name:'"+IG_datasets.cell(row=row,column=3).value+"'}) create (s)-[:ItemGroupRef]->(igd:ItemGroupDef {"
	firstproperty=True
	if IG_datasets.cell(row=row,column=1).value:
		for col in range(len(propertynames)):
			if IG_datasets.cell(row=row,column=col+1).value and col != 2:
				if not firstproperty:
					statement=statement+', '
				statement=statement+propertynames[col]+': "'+IG_datasets.cell(row=row,column=col+1).value+'" '
				firstproperty=False
		statement=statement+'}),(igd)-[:BasedOn]->(c)'
	print 'STATEMENT: '+statement
	tx.append(statement)

# Import variable level metadata and create the itemdef nodes
propertynames = ['Name', 'Label', 'SASType', 'Codelist', 'Core', 'DataType', 'MaxLength','Origin','Predecessor']
for row in range(2,IG_vars.max_row+1):
#for row in 3,135:
    statement = "merge (id:ItemDef {"
    firstproperty = True
    for col in range(len(propertynames)):
        if not firstproperty:
            statement=statement+', '
        statement=statement+propertynames[col]+': "'+str(IG_vars.cell(row=row,column=col+5).value)+'" '
        firstproperty = False

    # Create relationship to itemdef
    statement=statement+'}) on create set id.OID="ID.'+stdname+stdversion+'.'+str(IG_vars.cell(row=row,column=5).value)+'.'+str(row)+'" with id match (igd:ItemGroupDef {Name: "'+str(IG_vars.cell(row=row,column=3).value)+'"}) create (igd)-[:ItemRef'
    statement=statement+' {OrderNumber:'+str(IG_vars.cell(row=row,column=1).value)
    if IG_vars.cell(row=row,column=15).value:
        statement=statement+', MethodOID:"MT.'+modelname+modelversion+'.'+IG_vars.cell(row=row,column=4).value+'"'
    if IG_vars.cell(row=row,column=16).value:
        statement=statement+', Mandatory:"'+IG_vars.cell(row=row,column=16).value+'"'
    statement=statement+'}]->(id) '
    # Create methods and relationship to methods
    # if IG_vars.cell(row=row,column=15).value:
    #     statement=statement+' with id match (mt:Method {OID:"MT.'+modelname+modelversion+'.'+IG_vars.cell(row=row,column=4).value+'"}) merge (id)-[:MethodRef]->(mt) '
    #     tx.append(methodstmt)
    # Create relationship to dictionary
    if IG_vars.cell(row=row,column=14).value:
        statement=statement+' with id match (dict:Dictionary {Dictionary:"'+IG_vars.cell(row=row,column=14).value+'"}) merge (id)-[:DictionaryRef]->(dict) '
    # Create relationship to CT
    if IG_vars.cell(row=row,column=8).value:
        statement=statement+' with id match (ct:CodeList {OID:"'+IG_vars.cell(row=row,column=8).value+'"}) create (id)-[:CodeListRef]->(ct) '
    
    tx.append(statement)

tx.commit()