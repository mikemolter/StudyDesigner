# This program supercedes, as of today, 9/17/18, all other programs that create CT.
# It creates new CodeListSet nodes, one for SDTM and one for ADaM, each read from a different spreadsheet.
# After this, updated versions of CT will be created with a separate program, not yet written.
# Currently written for openpyxl 2.4.1

# Enter neo4j information here
neousername="neo4j"
neopassword="ne04j"
neopath="localhost:7474/db/data/"

#sdtmCTVersion='2018-06-29'
adamCTVersion='2017-09-29'
#wbsdtm='SDTM_Terminology_2018-06-29.xlsx'
wbadam='ADaM_Terminology_2017-09-29 (1).xlsx'
#sheetsdtm='SDTM Terminology 2018-06-29'
sheetadam='ADaM Terminology 2017-09-29'

import openpyxl
from py2neo import Graph
from py2neo.packages.httpstream import http
http.socket_timeout = 9999

# Open workbooks and worksheets
#wbs = openpyxl.load_workbook(wbsdtm)
#cts = wbs.get_sheet_by_name(sheetsdtm)
wba = openpyxl.load_workbook(wbadam)
cta = wba.get_sheet_by_name(sheetadam)
#ct  = [{'wb':cts,'version':sdtmCTVersion,'model':"SDTM"},{'wb':cta,'version':adamCTVersion,'model':'ADaM'}]
ct  = [{'wb':cta,'version':adamCTVersion,'model':'ADaM'}]

# graph = Graph('http://neo4j:letsgowings@10.0.0.10:7474/db/data/')
graph=Graph('http://'+neousername+':'+neopassword+'@'+neopath)
tx = graph.begin()

clipropnames=['AliasName','CodedValue','Decode']
clicolumns=[1,5,8]
clpropnames=['AliasName','Extensible','Name','OID']
clcolumns=[1,3,4,5]

for c in ct:
    statement='CREATE (ct:CodeListSet {Version: "'+c['version']+'",Model:"'+c['model']+'"}) '
    for row in range(2,c['wb'].max_row+1):
        firstproperty=True
        if c['wb'].cell(row=row,column=2).value:
            statement=statement+'with ct,cl MERGE (cli:CodeListItem {'
            for col in range(len(clipropnames)):
                if not firstproperty:
                    statement=statement+', '
                statement=statement+clipropnames[col]+': "'+str(c['wb'].cell(row=row,column=clicolumns[col]).value)+'" '
                firstproperty=False
            statement=statement+'}) WITH ct,cl,cli CREATE (cl)-[r:ContainsCodeListItem]->(cli) '
        else:
            statement=statement+'with ct CREATE (ct)-[:ContainsCodeList]->(cl:CodeList {DataType:"text"'
            for col in range(len(clpropnames)):
                statement=statement+', '+clpropnames[col]+': "'+str(c['wb'].cell(row=row,column=clcolumns[col]).value)+'"'
                firstproperty=False
            statement=statement+'}) '

    # Now add the YES codelist
    statement=statement+'with ct, cl,cli match (cl:CodeList {AliasName:"C66742"})--(cli:CodeListItem {CodedValue:"Y"}) create (new:CodeList {Name:"YES RESPONSE",OID:"YES",Extensible:"No",AliasName:"C66742",DataType:"text"})-[:ContainsCodeListItem]->(cli), \
        (new)-[:BasedOn]->(cl) '

    print ('STATEMENT: '+statement)
    tx.append(statement)
tx.commit()

# Changes to make
# 4) Add YESONLY BasedOn relationship with child=1
# 5) Use Decode flag to differentiate which codelists use enumerateditems and which don't
