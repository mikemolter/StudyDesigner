# This program imports the model from the spreadsheet
# It does not yet connect model variables to codelists.  That will come in a separate program.
# It also does not yet connect mdoel variables to predecessors.  That will come in a separate program too.

import openpyxl
from py2neo import Graph

modelname='ADaM'
modelversion='2.1'

wb = openpyxl.load_workbook('adam-2-1-draft.xlsx')
model_datasets = wb.get_sheet_by_name('ADaM 2.1 Dataset')
model_vars = wb.get_sheet_by_name('ADaM 2.1 Variable')

#graph = Graph("http://neo4j:letsgowings@10.0.0.10:7474/db/data/")
graph = Graph('http://neo4j:letsgowings@localhost:7474/db/data/')

tx = graph.cypher.begin()

# Create the model node
tx.append("CREATE (m:Model {name:'"+modelname+"', version:'"+modelversion+"'})")

# Create dictionary nodes (hard-code)
tx.append("create (:Dictionary {Dictionary:'WHO Drug',DataType:'text',Name:'World Health Organization Drug Dictionary'})")
tx.append("create (:Dictionary {Dictionary:'MedDRA',DataType:'text',Name:'Adverse Event Dictionary'})")

# Import dataset level metadata and create the itemgroupdef nodes
statement = "CREATE (igd:ItemGroupDef {Name: {C1}})"

for row in range(2,model_datasets.max_row+1):
    if model_datasets.cell(row=row,column=1).value:
        tx.append(statement,{"C1":model_datasets.cell(row=row,column=1).value})

# Create relationship from model node to itemgroupdefs
tx.append("MATCH (m:Model),(igd:ItemGroupDef) CREATE (m)-[r:ItemGroupRef]->(igd)")

# Import variable level metadata and create the itemdef nodes
propertynames = ['Name', 'ModelOID', 'Label', 'SASType', 'Codelist', 'Core', 'DataType', 'MaxLength','Origin','Predecessor']
for row in range(2,model_vars.max_row+1):
#for row in 8,208:
    statement = "merge (id:ItemDef {"
    firstproperty = True
    for col in range(len(propertynames)):
        if col != 1 and col != 4:
            if not firstproperty:
                statement=statement+', '
            statement=statement+propertynames[col]+': "'+str(model_vars.cell(row=row,column=col+4).value)+'" '
            firstproperty = False

    # Create an OID based on the Excel row, and only do so if a node with the same property values does not yet exist
    # Create relationship to itemdef
    statement=statement+'}) on create set id.OID="ID.'+modelname+modelversion+'.'+str(model_vars.cell(row=row,column=4).value)+"."+str(row)+'" with id match (igd:ItemGroupDef {Name: "'+str(model_vars.cell(row=row,column=2).value)+'"}) create (igd)-[:ItemRef'
    statement=statement+' {OrderNumber:'+str(model_vars.cell(row=row,column=1).value)
    if model_vars.cell(row=row,column=15).value:
        statement=statement+', MethodOID:"MT.'+modelname+modelversion+'.'+model_vars.cell(row=row,column=4).value+'"'
    if model_vars.cell(row=row,column=16).value:
        statement=statement+', Mandatory:"'+model_vars.cell(row=row,column=16).value+'"'
    statement=statement+'}]->(id) '
    # Create methods and relationship to methods
    if model_vars.cell(row=row,column=15).value:
        statement=statement+' with id match (mt:Method {OID:"MT.'+modelname+modelversion+'.'+model_vars.cell(row=row,column=4).value+'"}) create (id)-[:MethodRef]->(mt) '
        methodstmt='create (:Method {OID:"MT.'+modelname+modelversion+'.'+model_vars.cell(row=row,column=4).value+'", Type:"Computation", \
            Name: "Computation for '+model_vars.cell(row=row,column=6).value+'", Description: "'+model_vars.cell(row=row,column=15).value+'"}) '
        tx.append(methodstmt)
    # Create relationship to dictionary
    if model_vars.cell(row=row,column=14).value:
        statement=statement+' with id match (dict:Dictionary {Dictionary:"'+model_vars.cell(row=row,column=14).value+'"}) create (id)-[:DictionaryRef]->(dict) '
    # Create relationship to CT
    if model_vars.cell(row=row,column=8).value:
        statement=statement+' with id match (ct:CodeList {OID:"'+model_vars.cell(row=row,column=8).value+'"}) create (id)-[:CodeListRef]->(ct) '

    
    tx.append(statement)


tx.commit()
