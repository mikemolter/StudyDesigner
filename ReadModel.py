#!/usr/bin/python

import openpyxl
from py2neo import Graph

wb = openpyxl.load_workbook('Initial SDTM 1.4 IG 3.2.xlsx')
model_datasets = wb.get_sheet_by_name('SDTM 1.4 domains')
model_vars = wb.get_sheet_by_name('SDTM 1.4 Variables')
model_refs = wb.get_sheet_by_name('SDTM 1.4 itemrefs')
model_specs = wb.get_sheet_by_name('SDTM 1.4 Mapping Specs')

graph = Graph("http://neo4j:letsgowings@10.0.0.10:7474/db/data/")

tx = graph.cypher.begin()

tx.append("CREATE (m:Model {name:'SDTM', version:'1.4'})")

# Import dataset level metadata
statement = "CREATE (igd:ItemGroupDef {Name: {C1}, transpose: {C2}})"

for row in range(2,model_datasets.max_row+1):
  tx.append(statement,{"C1":model_datasets.cell(row=row,column=1).value, "C2":model_datasets.cell(row=row,column=2).value})

tx.append("MATCH (m:Model {name:'SDTM'}),(igd:ItemGroupDef) CREATE (m)-[r:ItemGroupRef]->(mc) RETURN r")

# Import variable level metadata
propertynames = ['Name', 'OID', 'Label', 'SASType', 'Length', 'DataType', 'Role', 'Codelist', 'Dictionary', 'Origin','VLMFlag']
for row in range(2,model_vars.max_row+1):
    statement = "CREATE (id:ItemDef {"
    firstproperty = True
    for col in range(len(propertynames)):
        if model_vars.cell(row=row,column=col+1).value:
            if col !=7 and col !=8:
                if not firstproperty:
                    statement=statement+', '
                statement=statement+propertynames[col]+': "'+str(model_vars.cell(row=row,column=col+1).value)+'" '
                firstproperty = False
    statement=statement+'})'
    tx.append(statement)

# Import relationships
propertynames = ['Order', 'VariableRequired', 'Mandatory', 'Key', 'MethodOID']
for row in range(2,model_refs.max_row+1):
    statement = "MATCH (n:ItemGroupDef {name: '"+model_refs.cell(row=row,column=1).value+"'}), (m:ItemDef {Name: '"+model_refs.cell(row=row,column=2).value+"'}) "
    statement = statement+"CREATE (n) - [r:ItemRef {"
    firstproperty = True
    for col in range(len(propertynames)):
        if model_refs.cell(row=row,column=col+4).value:
            if not firstproperty:
                statement=statement+', '
            statement=statement+propertynames[col]+': "'+str(model_refs.cell(row=row,column=col+4).value)+'" '
            firstproperty = False
    statement=statement+'}] -> (m)'
    tx.append(statement)

# Import Methods

for row in range(2,model_specs.max_row+1):
    statement = "CREATE (mc:MethodDef {MethodOID: {C1}, MapSpec: {C2}}) WITH mc "
    #Find the variable nodes that are to link to the Methods
    statement=statement+"MATCH ()-[r:ItemRef {MethodOID: {C1}}]->(n) "
    # Create the relationship from the variable node to the method node
    statement=statement+"MERGE (n)-[r1:MethodRef]->(mc)"

    tx.append(statement,{"C1":model_specs.cell(row=row,column=1).value, "C2":model_specs.cell(row=row,column=2).value})

# Create relationships to codelists
for row in range(2,model_vars.max_row+1):
    codelist=str(model_vars.cell(row=row,column=8).value)
    if codelist:
        statement='MATCH (ct:CodeList {OID : "'+codelist+'"}), (mv1:ItemDef {OID: "'+str(model_vars.cell(row=row,column=2).value)+'"}) MERGE (mv1)-[r1:CodelistRef]->(ct) '
        tx.append(statement)
        #print statement

tx.commit()
